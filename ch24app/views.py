import logging
from django.shortcuts import render, redirect, get_object_or_404
from .models import Creator, Program, Episode, EpisodeMediaInfo
from .forms import CreatorForm, ProgramForm, EpisodeForm, EpisodeUploadForm, EpisodeUpdateForm
from django.http import HttpResponseRedirect, HttpResponse
from .utils import create_presigned_url, convert_seconds_to_timecode  # Assuming the function is in utils.py
from django.contrib import messages
from django.urls import reverse
import requests
import boto3
from pymediainfo import MediaInfo
from .utils import validate_media_info, create_presigned_view_url  # Import the validation function
from django.core.exceptions import ValidationError

import os
from django.conf import settings
from .models import HomeMessage

AWS_STORAGE_BUCKET_NAME = settings.AWS_STORAGE_BUCKET_NAME

# from .models import MediaInfo

# logger = logging.getLogger('django')
# Get an instance of a logger
logger = logging.getLogger(__name__)

from django.shortcuts import render
from .models import StreamSettings
# Create your views here.
def home(request):
    user_has_creator = False
    user_has_programs = False
    if request.user.is_authenticated:
        user_has_creator = Creator.objects.filter(created_by=request.user).exists() 
        user_has_programs = Program.objects.filter(created_by=request.user).exists()
        print(f"user_has_creator: {user_has_creator}")
        print(f"user_has_programs: {user_has_programs}")

    active_message = HomeMessage.objects.filter(is_active=True).first()

    stream_settings, created = StreamSettings.objects.get_or_create(
        defaults={'is_stream_active': True}
    )

    context = {
        'show_stream': stream_settings.is_stream_active,
        'user_has_creator': user_has_creator, 
        'user_has_programs': user_has_programs,
        'active_message': active_message
    }

    return render(request, 'home.html', context)

def all_creators(request):
    creator_list = Creator.objects.all()
    return render(request, 'creator_list.html', {'creator_list': creator_list})

def all_programs(request):
    program_list = Program.objects.all()

    program_list = program_list.annotate(
        episode_count=models.Count('episode')
    )
    
    return render(request, 'program_list.html', {'program_list': program_list})

def all_episodes(request):
    episode_list = Episode.objects.all()
    return render(request, 'episode_list.html', {'episode_list': episode_list})

def my_creators(request):
    creator_list = Creator.objects.filter(created_by=request.user)
    return render(request, 'my_creators.html', {'creator_list': creator_list})

# def my_programs(request):
#     program_list = Program.objects.filter(created_by=request.user)
#     return render(request, 'my_programs.html', {'program_list': program_list})

# from django.db.models import Count
from django.db import models

def my_programs(request):
    # Get all programs for the user
    program_list = Program.objects.filter(created_by=request.user)
    
    # For each program, annotate with episode count
    program_list = program_list.annotate(
        episode_count=models.Count('episode')
    )
    
    return render(request, 'my_programs.html', {'program_list': program_list})

def my_episodes(request):
    episode_list = Episode.objects.filter(created_by=request.user)
    return render(request, 'my_episodes.html', {'episode_list': episode_list})


def add_creator(request):
    submitted = False
    form = CreatorForm()
    if request.method == 'POST':
        form = CreatorForm(request.POST, user=request.user)
        if form.is_valid():
            instance = form.save(commit=False)  # Create an instance without saving to the database
            instance.created_by = request.user  # Set the created_by field to the current user
            instance.save()  # Now save the instance to the database
            return HttpResponseRedirect('/add_creator?submitted=True')
    else:
        form = CreatorForm(user=request.user)
        if 'submitted' in request.GET:
            submitted = True
    return render(request, 'add_creator.html', {'form': form, 'submitted': submitted})


# views.py

from django.shortcuts import render, HttpResponseRedirect
from .forms import ProgramForm
from .models import Creator


from django.urls import reverse

# def add_program(request):
#     submitted = False
#     if request.method == 'POST':
#         form = ProgramForm(request.POST, user=request.user)
#         if form.is_valid():
#             instance = form.save(commit=False)
#             instance.created_by = request.user
#             instance.save()
#             return HttpResponseRedirect(f'{reverse("add-program")}?submitted=True')
#     else:
#         initial_data = {}
#         last_creator = Creator.objects.filter(created_by=request.user).order_by('-created_at').first()
#         if not last_creator:
#             messages.error(request, "You need to add a Channel before adding programs.")
#             creator_form = CreatorForm(user=request.user)
#             submitted = False
#             return render(request, 'add_creator.html', {'form': creator_form, 'submitted': submitted})        

#         if last_creator:
#             initial_data['creator'] = last_creator  # Pre-fill the 'program' field

#         form = ProgramForm(user=request.user, initial=initial_data)
#         if 'submitted' in request.GET:
#             submitted = True
#     return render(request, 'add_program.html', {'form': form, 'submitted': submitted})

def add_program(request):
    submitted = False
    if request.method == 'POST':
        # Debugging
        print("Original POST data:", request.POST)
        print("time_slots_requested in POST:", request.POST.getlist('time_slots_requested'))
        # Get the POST data
        post_data = request.POST.copy()
        
        # Check if time_slots_requested is in the POST data and is a string that looks like a list
        if 'time_slots_requested' in post_data and post_data['time_slots_requested'].startswith('['):
            # Convert the string representation of a list to an actual list
            import ast
            try:
                # Safely evaluate the string as a Python literal
                time_slots_list = ast.literal_eval(post_data['time_slots_requested'])
                # Replace the string with multiple values
                post_data.setlist('time_slots_requested', time_slots_list)
            except (ValueError, SyntaxError):
                # If there's an error parsing, just continue with the original data
                pass
                
        form = ProgramForm(post_data, user=request.user)
        # form = ProgramForm(request.POST, user=request.user)
        print("Form errors:", form.errors)
        
        if form.is_valid():
            instance = form.save(commit=False)
            instance.created_by = request.user
            instance.save()
            return HttpResponseRedirect(f'{reverse("add-program")}?submitted=True')
    else:
        initial_data = {}
        last_creator = Creator.objects.filter(created_by=request.user).order_by('-created_at').first()
        if not last_creator:
            messages.error(request, "You need to add a Channel before adding programs.")
            creator_form = CreatorForm(user=request.user)
            submitted = False
            return render(request, 'add_creator.html', {'form': creator_form, 'submitted': submitted})        

        if last_creator:
            initial_data['creator'] = last_creator  # Pre-fill the 'creator' field

        form = ProgramForm(user=request.user, initial=initial_data)
        if 'submitted' in request.GET:
            submitted = True
    return render(request, 'add_program.html', {'form': form, 'submitted': submitted})



def no_creator_found(request):
    return render(request, 'no_creator_found.html')


def add_episode(request):
    if request.method == 'POST':
        form = EpisodeForm(request.POST, user=request.user)
        if form.is_valid():
            episode = form.save(commit=False)
            episode.created_by = request.user
            episode.save()
            # Redirect with episode_id in query parameters
            return HttpResponseRedirect(f'{reverse("add-episode")}?submitted=True&episode_id={episode.custom_id}')
    else:
        # Prepare initial data for the form
        initial_data = {}
        last_program = Program.objects.filter(created_by=request.user).order_by('-created_at').first()
        if not last_program:
            messages.error(request, "You need to add a program before adding episodes.")
            program_form = ProgramForm(user=request.user)
            return render(request, 'add_program.html', {'form': program_form, 'submitted': False})        

        if last_program:
            initial_data['program'] = last_program  # Pre-fill the 'program' field

        form = EpisodeForm(user=request.user, initial=initial_data)
    
    submitted = 'submitted' in request.GET
    episode = None
    if submitted and 'episode_id' in request.GET:
        episode_id = request.GET.get('episode_id')
        episode = get_object_or_404(Episode, custom_id=episode_id)
    
    return render(request, 'add_episode.html', {'form': form, 'submitted': submitted, 'episode': episode})

# views.py

from django.shortcuts import render, get_object_or_404
from .models import Episode
from .forms import EpisodeAnalysisForm

def episode_analysis_view(request, custom_id):
    """
    View to display the analysis of a specific episode in read-only mode.
    """
    # Fetch the Episode instance based on the custom_id
    episode = get_object_or_404(Episode, custom_id=custom_id)
    
    # Instantiate the form with the Episode instance
    form = EpisodeAnalysisForm(instance=episode)
    
    context = {
        'form': form,
        'episode': episode,  # Optional: Pass the episode object if needed in the template
    }
    
    return render(request, 'my_episode_analysis.html', context)


def update_creator(request, creator_id):
    creator = Creator.objects.get(custom_id=creator_id)  # Changed from id to custom_id
    if request.method == "POST":
        form = CreatorForm(request.POST, instance=creator)
        if form.is_valid():
            form.save()
            return render(request, 'update_creator.html', {
                'form': form,
                'submitted': True
            })
    else:
        form = CreatorForm(instance=creator)
    
    return render(request, 'update_creator.html', {
        'form': form,
        'submitted': False
    })

from django.shortcuts import render, get_object_or_404
from .models import Episode, TIME_SLOTS_CHOICES
# from .utils import TIME_SLOTS
from .forms import EpisodeAnalysisForm

from django.shortcuts import render, get_object_or_404
from .forms import EpisodeAnalysisForm
from .models import Episode

def update_analysis(request, custom_id):
    episode = get_object_or_404(Episode, custom_id=custom_id)

    if request.method == "POST":
        form = EpisodeAnalysisForm(request.POST, instance=episode)
        if form.is_valid():
            print("Form is valid!")
            form.save()
            return render(request, 'update_analysis.html', {'form': form, 'submitted': True})
        else:
            print(form.errors)

    else:
        form = EpisodeAnalysisForm(instance=episode)

    return render(request, 'update_analysis.html', {'form': form, 'submitted': False})


# views.py

from django.urls import reverse
from django.http import HttpResponseRedirect

def update_program(request, program_id):
    program = Program.objects.get(pk=program_id)
    if request.method == 'POST':
        form = ProgramForm(request.POST, instance=program, user=request.user)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(f'{reverse("update-program", args=[program_id])}?submitted=True')
    else:
        form = ProgramForm(instance=program, user=request.user)
    return render(request, 'update_program.html', {'form': form, 'submitted': request.GET.get('submitted', False)})



def update_episode(request, episode_id):
    episode = Episode.objects.get(custom_id=episode_id)  # Changed from id to custom_id
    if request.method == "POST":
        form = EpisodeUpdateForm(request.POST, instance=episode)
        if form.is_valid():
            form.save()
            return render(request, 'update_episode.html', {
                'form': form,
                'submitted': True
            })
    else:
        form = EpisodeUpdateForm(instance=episode)
    
    return render(request, 'update_episode.html', {
        'form': form,
        'submitted': False
    })

def get_mediainfo_from_s3(bucket_name, s3_key):
    # Initialize a session using Amazon S3
    s3 = boto3.client('s3')

    # Define a temporary file path
    temp_file_path = os.path.join(settings.MEDIA_ROOT, '', os.path.basename(s3_key))

    print('Temp file path: ', temp_file_path)
    logger.info(f'Temp file path: {temp_file_path}')

    # Download the file from S3
    s3.download_file(bucket_name, s3_key, temp_file_path)

    # Use pymediainfo to get media information
    media_info = MediaInfo.parse(temp_file_path)
    # pprint(media_info)

    # Clean up the temporary file
    os.remove(temp_file_path)

    return media_info


# views.py

from .utils import validate_media_info  # Ensure this is imported

import boto3
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from django.urls import reverse
from django.conf import settings

from .forms import EpisodeUploadForm
from .models import Episode, EpisodeMediaInfo
# from .mediainfo_utils import get_mediainfo_from_s3, validate_media_info

AWS_STORAGE_BUCKET_NAME = getattr(settings, 'AWS_STORAGE_BUCKET_NAME', None)

def sanitize_filename(filename):
    """
    Remove all characters outside of ASCII range.
    You can make this more selective as needed
    (e.g., only remove specific emoji ranges).
    """
    # Encode to ASCII and ignore errors, then decode back to str
    return filename.encode('ascii', errors='ignore').decode()

def verify_s3_object(bucket_name, object_key):
    """
    Check if an object with the specified key exists in S3.
    Returns True if it exists, False otherwise.
    """
    s3_client = boto3.client('s3')
    try:
        s3_client.head_object(Bucket=bucket_name, Key=object_key)
        return True
    except Exception:
        return False

def upload_episode(request, episode_id):
    episode = get_object_or_404(Episode, custom_id=episode_id)

    # Security check
    if episode.created_by != request.user:
        return HttpResponse("Unauthorized", status=401)

    if request.method == 'POST':
        form = EpisodeUploadForm(request.POST, request.FILES)
        if form.is_valid():
            # Delete previous EpisodeMediaInfo instances
            EpisodeMediaInfo.objects.filter(episode=episode).delete()

            # Original filename from the uploaded file
            file = form.cleaned_data['file']
            original_file_name = file.name

            # Sanitize filename to remove emojis/unwanted unicode
            sanitized_file_name = sanitize_filename(original_file_name)

            # Construct S3 path: e.g. "episode123/myfile.mp4"
            bucket_name = AWS_STORAGE_BUCKET_NAME
            unique_file_name = f'{episode.custom_id}/{sanitized_file_name}'

            # Initialize the S3 client
            s3_client = boto3.client('s3')

            # Upload the file to S3
            try:
                with file.open('rb') as f:
                    s3_client.upload_fileobj(f, bucket_name, unique_file_name)
            except Exception as e:
                messages.error(request, f"Upload failed: {e}")
                return redirect('upload_failed')

            # Verify the object is actually in S3 after upload
            if not verify_s3_object(bucket_name, unique_file_name):
                messages.error(request, 
                    f"Upload failed: Could not verify '{unique_file_name}' in S3.")
                return redirect('upload_failed')

            # Optional message if sanitization changed the file name
            if original_file_name != sanitized_file_name:
                messages.info(
                    request,
                    f"Filename changed from '{original_file_name}' "
                    f"to '{sanitized_file_name}' to remove invalid characters."
                )

            # Retrieve the media info from the uploaded file
            media_info = get_mediainfo_from_s3(bucket_name, unique_file_name)

            # Save the media info to the database
            track_id = 0
            for track in media_info.tracks:
                track_id += 1
                track_metadata = {
                    key: value for key, value in track.to_data().items() if value is not None
                }
                EpisodeMediaInfo.objects.create(
                    episode=episode,
                    track_id=track_id,
                    metadata=track_metadata
                )

            # Validate media info
            media_infos = episode.media_infos.all()
            unique_errors, unique_warnings = validate_media_info(media_infos)
            episode.has_mediainfo_errors = bool(unique_errors)

            # IMPORTANT: store the sanitized file name in the DB
            episode.file_name = unique_file_name
            episode.save()

            # Attempt to populate episode duration fields
            try:
                media_info_track1 = EpisodeMediaInfo.objects.get(episode=episode, track_id=1)
                duration_ms = media_info_track1.metadata.get('duration')
                if duration_ms:
                    duration_sec = float(duration_ms) / 1000
                    episode.duration_seconds = duration_sec
                    episode.duration_timecode = convert_seconds_to_timecode(duration_sec)
                    episode.save()
            except EpisodeMediaInfo.DoesNotExist:
                pass

            messages.success(request, "File uploaded successfully.")
            return redirect(f"{reverse('upload_success')}?episode_id={episode.custom_id}")

        else:
            messages.error(request, "Form is invalid.")
    else:
        form = EpisodeUploadForm()

    return render(request, 'episode_upload.html', {'form': form, 'episode': episode})


def upload_success(request):
    episode_id = request.GET.get('episode_id')
    episode = get_object_or_404(Episode, custom_id=episode_id) if episode_id else None

    return render(request, 'upload_success.html', {
        'episode': episode,
    })



def upload_failed(request):
    return render(request, 'upload_failed.html')

def adobe_premiere(request):
    return render(request, 'adobe_premiere.html')

def davinci_resolve(request):
    return render(request, 'davinci_resolve.html')

# def getting_started2(request):
#     return render(request, 'getting_started2.html')

def getting_started(request):
    return render(request, 'getting_started.html')

def episode_media_info(request, episode_id):
    episode = get_object_or_404(Episode, custom_id=episode_id)
    media_infos = episode.media_infos.all()

    # Use the validation function
    unique_errors, unique_warnings = validate_media_info(media_infos)

    # Update the has_mediainfo_errors field
    episode.has_mediainfo_errors = bool(unique_errors)
    episode.save()

    # Add unique error messages to the Django messages framework as errors
    for message in unique_errors:
        messages.error(request, message)

    # Add unique warning messages to the Django messages framework as warnings
    for message in unique_warnings:
        messages.warning(request, message)

    return render(request, 'episode_media_info.html', {
            'episode': episode,
            'media_infos': media_infos,
            'errors': unique_errors,
            'warnings': unique_warnings
        })


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from .forms import SupportTicketForm, SupportTicketStatusForm, TicketResponseForm
from .models import SupportTicket, TicketResponse
from .notifications import send_ticket_notification

@login_required
def submit_ticket(request):
    if request.method == 'POST':
        form = SupportTicketForm(request.POST, user=request.user)
        if form.is_valid():
            ticket = form.save(commit=False)
            if request.user.is_authenticated:
                selected_creator = form.cleaned_data.get('creator')
                if selected_creator:
                    ticket.creator = selected_creator
                else:
                    user_creator = Creator.objects.filter(created_by=request.user).first()
                    if user_creator:
                        ticket.creator = user_creator
                ticket.created_by = request.user
            ticket.save()
            
            # Send notification email
            try:
                send_ticket_notification(ticket)
            except Exception as e:
                # Log the error but don't prevent ticket creation
                logger = logging.getLogger('django.request')
                logger.error(f'Failed to send ticket notification email: {str(e)}')
            
            return redirect('ticket_submitted', ticket_no=ticket.ticket_no)
    else:
        form = SupportTicketForm(user=request.user)
    return render(request, 'support/submit_ticket.html', {'form': form})

def ticket_submitted(request, ticket_no):
    ticket = get_object_or_404(SupportTicket, ticket_no=ticket_no)
    return render(request, 'support/ticket_submitted.html', {'ticket': ticket})

def ticket_detail(request, ticket_no):
    ticket = get_object_or_404(SupportTicket, ticket_no=ticket_no)
    responses = ticket.responses.all().order_by('response_no')
    status_form = None
    response_form = TicketResponseForm()

    if request.method == 'POST':
        if 'status_form' in request.POST and request.user.is_staff:
            status_form = SupportTicketStatusForm(request.POST, instance=ticket)
            if status_form.is_valid():
                status_form.save()
                return redirect('ticket_detail', ticket_no=ticket_no)
        elif 'response_form' in request.POST:
            response_form = TicketResponseForm(request.POST)
            if response_form.is_valid():
                response = response_form.save(commit=False)
                response.ticket = ticket
                if request.user.is_authenticated:
                    response.responder = request.user
                response.save()
                return redirect('ticket_detail', ticket_no=ticket_no)
    else:
        if request.user.is_staff:
            status_form = SupportTicketStatusForm(instance=ticket)

    return render(request, 'support/ticket_detail.html', {
        'ticket': ticket,
        'responses': responses,
        'response_form': response_form,
        'status_form': status_form,
    })


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import SupportTicket

@login_required
def my_tickets(request):
    tickets = SupportTicket.objects.filter(created_by=request.user).order_by('-time_received')
    return render(request, 'support/my_tickets.html', {'tickets': tickets})

@login_required
def view_episode(request, episode_id):
    episode = get_object_or_404(Episode, custom_id=episode_id)

    # Security check: Ensure the user is allowed to view the episode
    # if episode.created_by != request.user:
    #     return HttpResponse("Unauthorized", status=401)

    bucket_name = AWS_STORAGE_BUCKET_NAME
    s3_key = episode.file_name

    # Generate pre-signed URL
    presigned_url = create_presigned_view_url(bucket_name, s3_key)

    if presigned_url:
        # Pass the URL to the template
        return render(request, 'view_episode.html', {
            'episode': episode,
            'presigned_url': presigned_url,
        })
    else:
        messages.error(request, "Unable to generate view URL.")
        return redirect('available_content')  # Replace with your actual error page


from django.http import HttpResponse

def health_check(request):
    return HttpResponse("OK", status=200)

import os
from django.http import JsonResponse

def environment(request):
    """
    View to display all environment variables.
    """
    excluded_keywords = ["DATABASE", "DJANGO", "AWS", "PYENV", "HOME", "USER", "PATH", "LOGNAME", "PWD", "VIRTUAL_ENV", "ZDOTDIR"]  # Keywords to exclude
    env_vars = {
        key: value
        for key, value in os.environ.items()
        if not any(keyword in key for keyword in excluded_keywords)
    }
    return render(request, 'environment_variables.html', {'env_vars': env_vars})


# views.py
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.contrib import messages
from datetime import datetime
from .utils import schedule_episodes
from django.utils import timezone
from .models import Episode, ScheduledEpisode
import csv
import io
import pytz



# views.py
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils import timezone
from django.contrib import messages
from datetime import datetime
import pytz

@login_required
@user_passes_test(lambda u: u.is_staff)
def playlist_create(request):
    logger.debug("Entered playlist_create view. Request method: %s", request.method)

    if request.method == 'POST':
        action = request.POST.get('action', '')
        playlist_date_str = request.POST.get('playlist_date', '')

        # Log the form submission details
        logger.debug("Form POST data -> action: '%s', playlist_date: '%s'", action, playlist_date_str)

        if not playlist_date_str:
            messages.error(request, 'Please select a date')
            logger.debug("No date provided in POST. Returning with error message.")
            return render(request, 'playlist_create.html')

        try:
            playlist_date = datetime.strptime(playlist_date_str, '%Y-%m-%d').date()
            logger.debug("Parsed playlist_date as: %s", playlist_date)

            if action == 'create':
                logger.debug("Checking if schedule already exists for %s", playlist_date)
                existing_schedule = ScheduledEpisode.objects.filter(schedule_date=playlist_date).exists()
                logger.debug("existing_schedule = %s", existing_schedule)

                if existing_schedule:
                    messages.warning(request, f'Schedule already exists for {playlist_date}. Clear it first.')
                    logger.debug("Schedule already exists. Skipping creation.")
                else:
                    logger.debug("Calling schedule_episodes() for date %s", playlist_date)
                    schedule_episodes(playlist_date, all_ready=False)
                    messages.success(request, f'Playlist created for {playlist_date}')

            elif action == 'clear':
                logger.debug("Clearing scheduled episodes for date %s", playlist_date)
                deleted_count = ScheduledEpisode.objects.filter(schedule_date=playlist_date).delete()[0]
                messages.success(request, f'Cleared {deleted_count} scheduled episodes for {playlist_date}!')
                logger.debug("Deleted %s scheduled episodes for date %s.", deleted_count, playlist_date)

            elif action == 'export':
                logger.debug("Exporting playlist for %s", playlist_date)
                return export_playlist(playlist_date)

            else:
                # If the action isn't recognized, it's probably just the date auto-submit
                logger.debug("No recognized action. Possibly the form auto-submitted on date change.")

        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            logger.exception("An error occurred in playlist_create view.")

    # Instead of always using 'today', parse the selected date from POST or GET.
    selected_date_str = request.POST.get('playlist_date') or request.GET.get('playlist_date')
    if selected_date_str:
        try:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except ValueError:
            # Fallback if parsing fails
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()

    # Query the schedule using the selected date
    scheduled_episodes = ScheduledEpisode.objects.filter(
        schedule_date=selected_date
    ).select_related('episode', 'program', 'creator').order_by('start_time')

    context = {
        'scheduled_episodes': scheduled_episodes,
        'selected_date': selected_date.strftime('%Y-%m-%d'),
    }

    logger.debug(
        "Rendering playlist_create.html with %s scheduled episodes for date: %s",
        scheduled_episodes.count(),
        selected_date
    )

    return render(request, 'playlist_create.html', context)


from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from datetime import datetime
import io
import pytz
import csv

# @login_required
# @user_passes_test(lambda u: u.is_staff)
def export_playlist(schedule_date):
    """
    Export a .ply playlist with lines like:
    "Z:\4bf15aac-059a-4de5-b17f-557703ace10a\ID1_2997_10Mbps.mp4";0.00000;41.07400;;ID1_2997_10Mbps
    """

    # If schedule_date is a string like "2025-01-08", parse it
    if isinstance(schedule_date, str):
        schedule_date = datetime.strptime(schedule_date, '%Y-%m-%d').date()

    # Prepare StringIO as our in-memory text buffer
    buffer = io.StringIO()

    # Query all ScheduledEpisode for the given date
    schedule = ScheduledEpisode.objects.filter(
        schedule_date=schedule_date
    ).order_by('start_time')

    # Write each line according to your requested format
    # Example line:
    # "Z:\{episode.custom_id}\{episode.file_name}";0.00000;{duration};;{title}
    for episode in schedule:
        # Build the file path in quotes
        # file_path = f"\"Z:\\{episode.custom_id}\\{episode.file_name}\""
        file_path = f"\"Z:\\{episode.file_name}\"".replace('/', '\\')

        # Start time is always 0.00000 in your example
        start_str = "0.00000"

        # Format duration with five decimals.
        # If duration_seconds is None, default to zero
        duration_value = float(episode.duration_seconds or 0)
        duration_str = f"{duration_value:.5f}"

        # Title (no quotes in your sample, just appended)
        title_str = episode.title

        # Construct the line
        line = f"{file_path};{start_str};{duration_str};;{title_str}"

        # Write the line plus a newline
        buffer.write(line + "\n")

    # Reset buffer pointer to the beginning
    buffer.seek(0)

    # Build the HTTP response for download
    response = HttpResponse(
        buffer,
        content_type='text/plain'  # or 'text/csv', but .ply is not a standard mime type
    )
    # Output file named with .ply extension
    filename = f"{schedule_date.strftime('%Y_%m_%d')}_00_00_00.ply"

    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



# views.py
from django.shortcuts import render, redirect
from .forms import ScheduledEpisodeForm
from django.contrib import messages

@login_required
@user_passes_test(lambda u: u.is_staff)
def scheduled_episode_create(request):
    if request.method == 'POST':
        form = ScheduledEpisodeForm(request.POST)
        if form.is_valid():
            try:
                scheduled_episode = form.save()
                messages.success(request, 'Episode scheduled successfully')
                return redirect('scheduled_episode_detail', pk=scheduled_episode.pk)
            except ValidationError as e:
                messages.error(request, str(e))
    else:
        form = ScheduledEpisodeForm()
    
    return render(request, 'scheduled_episode_form.html', {'form': form})

from django.views.generic import ListView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import JsonResponse
from django.db.models import Q
from django.shortcuts import get_object_or_404
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_POST
from .models import Episode, GENRE_CHOICES, AGE_RATING_CHOICES, Creator
import logging

logger = logging.getLogger(__name__)

class AvailableContentView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Episode
    template_name = 'available_content.html'
    context_object_name = 'episodes'
    paginate_by = 20

    def test_func(self):
        return self.request.user.is_staff
    
    @method_decorator(require_POST)
    def post(self, request, *args, **kwargs):
        # user must already be staff by the time this is reached
        episode_id = request.POST.get('episode_id')
        field = request.POST.get('field')
        value = request.POST.get('value')

        logger.info(f"Updating Episode: {episode_id}, Field: {field}, Value: {value}")

        try:
            episode = get_object_or_404(Episode, custom_id=episode_id)
            if field == 'ready_for_air':
                episode.ready_for_air = (value.lower() == 'true')
            elif field == 'ai_age_rating':
                episode.ai_age_rating = value
            # Add more fields as needed

            episode.save()
            return JsonResponse({'status': 'success', 'new_value': getattr(episode, field)})

        except Exception as e:
            logger.error(f"Error updating episode {episode_id}: {str(e)}", exc_info=True)
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    def get_queryset(self):
        # Normal GET filtering logic
        queryset = Episode.objects.all()

        # Clean all GET parameters by stripping whitespace
        clean_get = {}
        for key, value in self.request.GET.items():
            if isinstance(value, str):
                clean_get[key] = value.strip()
            else:
                clean_get[key] = value

        ready_for_air = clean_get.get('ready_for_air')
        if ready_for_air:
            if ready_for_air.lower() == 'true':
                queryset = queryset.filter(ready_for_air=True)
            elif ready_for_air.lower() == 'false':
                queryset = queryset.filter(ready_for_air=False)

        ai_genre = clean_get.get('ai_genre')
        if ai_genre:
            queryset = queryset.filter(ai_genre=ai_genre)

        ai_age_rating = clean_get.get('ai_age_rating')
        if ai_age_rating:
            queryset = queryset.filter(ai_age_rating=ai_age_rating)

        creator = clean_get.get('creator')
        if creator:
            queryset = queryset.filter(program__creator=creator)

        duration = clean_get.get('duration')
        if duration == 'bumper':
            queryset = queryset.filter(duration_seconds__lte=15)
        elif duration == 'shortform':
            queryset = queryset.filter(duration_seconds__gt=15, duration_seconds__lte=900)
        elif duration == 'longform':
            queryset = queryset.filter(duration_seconds__gt=900)

        search = clean_get.get('search')
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search)
                | Q(program__program_name__icontains=search)
                | Q(program__creator__channel_name__icontains=search)
            )

        sort = clean_get.get('sort', '-created_at')
        # Optional: Validate sort to avoid malicious input
        queryset = queryset.order_by(sort)

        return queryset.select_related('program', 'program__creator')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Clean all GET parameters
        clean_get = {}
        for key, value in self.request.GET.items():
            if isinstance(value, str):
                clean_get[key] = value.strip()
            else:
                clean_get[key] = value
        
        context['current_filters'] = {
            'ai_genre': clean_get.get('ai_genre', ''),
            'ai_age_rating': clean_get.get('ai_age_rating', ''),
            'creator': clean_get.get('creator', ''),
            'duration': clean_get.get('duration', ''),
            'search': clean_get.get('search', ''),
            'sort': clean_get.get('sort', '-created_at'),
            'ready_for_air': clean_get.get('ready_for_air', ''),
        }
        context['ai_genres'] = dict(GENRE_CHOICES)
        context['ai_age_ratings'] = dict(AGE_RATING_CHOICES)
        context['creators'] = Creator.objects.all()
        context['duration_choices'] = [
            ('bumper', 'Bumper (≤15s)'),
            ('shortform', 'Short Form (15s-15m)'),
            ('longform', 'Long Form (>15m)'),
        ]
        context['ready_for_air_choices'] = [
            ('', 'All Content'),
            ('true', 'Ready for Air'),
            ('false', 'Not Ready'),
        ]
        
        # Handle out-of-range page numbers gracefully
        if hasattr(context, 'paginator') and hasattr(context, 'page_obj'):
            paginator = context['paginator']
            page_number = self.request.GET.get('page', '1').strip()
            
            try:
                page_number = int(page_number)
                if page_number > paginator.num_pages and paginator.num_pages > 0:
                    # If requested page is out of range, deliver last page
                    context['page_obj'] = paginator.page(paginator.num_pages)
                    context['object_list'] = context['page_obj'].object_list
            except (ValueError, TypeError):
                # If page is not an integer, deliver first page
                pass
                
        return context
    
    def get(self, request, *args, **kwargs):
        # Clean the page parameter
        if 'page' in request.GET:
            page = request.GET.get('page', '1').strip()
            # Create a copy to modify
            get_copy = request.GET.copy()
            get_copy['page'] = page
            request.GET = get_copy
            
        return super().get(request, *args, **kwargs)        

# views.py
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Q
from django.core.paginator import Paginator

@staff_member_required
def admin_tickets(request):
    # Get filter parameters from request
    status = request.GET.get('status', '')
    category = request.GET.get('category', '')
    search = request.GET.get('search', '')
    urgency = request.GET.get('urgency', '')

    # Start with all tickets
    tickets = SupportTicket.objects.all().order_by('-time_received')

    # Apply filters
    if status:
        tickets = tickets.filter(ticket_status=status)
    if category:
        tickets = tickets.filter(category=category)
    if urgency:
        tickets = tickets.filter(urgency=urgency)
    if search:
        tickets = tickets.filter(
            Q(ticket_no__icontains=search) |
            Q(subject__icontains=search) |
            Q(name__icontains=search) |
            Q(contact_info__icontains=search)
        )

    # Pagination
    paginator = Paginator(tickets, 25)  # Show 25 tickets per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'tickets': page_obj,
        'status_choices': SupportTicket.STATUS_CHOICES,
        'category_choices': SupportTicket.TICKET_CATEGORIES,
        'urgency_choices': SupportTicket.URGENCY_CHOICES,
        'selected_status': status,
        'selected_category': category,
        'selected_urgency': urgency,
        'search_query': search,
    }
    
    return render(request, 'support/admin_tickets.html', context)


# views.py
@staff_member_required
def update_ticket_status(request, ticket_no):
    if request.method == 'POST':
        ticket = get_object_or_404(SupportTicket, ticket_no=ticket_no)
        new_status = request.POST.get('ticket_status')
        if new_status in dict(SupportTicket.STATUS_CHOICES):
            ticket.ticket_status = new_status
            ticket.save()
    return redirect('admin_tickets')


@login_required
def delete_episode(request, episode_id):
    logger.info(f"Delete episode request received for episode_id: {episode_id}")
    
    try:
        episode = get_object_or_404(Episode, custom_id=episode_id)
        logger.info(f"Episode found: {episode.title} (ID: {episode.custom_id})")
        
        # Security check: ensure the user owns this episode
        if episode.created_by != request.user:
            logger.warning(f"Unauthorized delete attempt for episode {episode_id} by user {request.user}")
            return HttpResponse("Unauthorized", status=401)
        
        if request.method == 'POST':
            logger.info(f"Processing POST request to delete episode {episode_id}")
            
            # If the episode has a file, delete it from S3
            if episode.file_name:
                try:
                    logger.info(f"Attempting to delete S3 file: {episode.file_name}")
                    s3_client = boto3.client('s3')
                    s3_client.delete_object(
                        Bucket=AWS_STORAGE_BUCKET_NAME,
                        Key=episode.file_name
                    )
                    logger.info("S3 file deleted successfully")
                except Exception as e:
                    logger.error(f"Error deleting S3 file: {str(e)}", exc_info=True)
                    messages.error(request, f"Error deleting file from S3: {str(e)}")
                    return redirect('my-episodes')
            
            try:
                # Delete any associated media info records first
                media_info_count = EpisodeMediaInfo.objects.filter(episode=episode).count()
                logger.info(f"Found {media_info_count} media info records to delete")
                EpisodeMediaInfo.objects.filter(episode=episode).delete()
                logger.info("Media info records deleted successfully")
                
                # Delete the episode
                logger.info(f"Attempting to delete episode {episode_id} from database")
                episode.delete()
                logger.info(f"Episode {episode_id} successfully deleted from database")
                
                messages.success(request, "Episode successfully deleted.")
                return redirect('my-episodes')
            except Exception as e:
                logger.error(f"Error deleting episode from database: {str(e)}", exc_info=True)
                messages.error(request, f"Error deleting episode: {str(e)}")
                return redirect('my-episodes')
        else:
            logger.warning(f"Non-POST request received for episode deletion: {request.method}")
        
        return redirect('my-episodes')
        
    except Episode.DoesNotExist:
        logger.error(f"Episode not found: {episode_id}")
        messages.error(request, "Episode not found.")
        return redirect('my-episodes')
    except Exception as e:
        logger.error(f"Unexpected error in delete_episode: {str(e)}", exc_info=True)
        messages.error(request, f"An unexpected error occurred: {str(e)}")
        return redirect('my-episodes')
    
@login_required
def delete_program(request, program_id):
    logger.info(f"Delete program request received for program_id: {program_id}")
    
    try:
        program = get_object_or_404(Program, custom_id=program_id)
        logger.info(f"Program found: {program.program_name} (ID: {program.custom_id})")
        
        # Security check: ensure the user owns this program
        if program.created_by != request.user:
            logger.warning(f"Unauthorized delete attempt for program {program_id} by user {request.user}")
            return HttpResponse("Unauthorized", status=401)
        
        # Check if program has any episodes
        episode_count = Episode.objects.filter(program=program).count()
        if episode_count > 0:
            logger.warning(f"Attempted to delete program {program_id} with {episode_count} episodes")
            messages.error(request, "Cannot delete program with existing episodes")
            return redirect('my-programs')
        
        if request.method == 'POST':
            logger.info(f"Processing POST request to delete program {program_id}")
            
            try:
                # Delete the program
                program.delete()
                logger.info(f"Program {program_id} successfully deleted from database")
                messages.success(request, "Program successfully deleted.")
                return redirect('my-programs')
            except Exception as e:
                logger.error(f"Error deleting program from database: {str(e)}", exc_info=True)
                messages.error(request, f"Error deleting program: {str(e)}")
                return redirect('my-programs')
        
        return redirect('my-programs')
        
    except Program.DoesNotExist:
        logger.error(f"Program not found: {program_id}")
        messages.error(request, "Program not found.")
        return redirect('my-programs')
    except Exception as e:
        logger.error(f"Unexpected error in delete_program: {str(e)}", exc_info=True)
        messages.error(request, f"An unexpected error occurred: {str(e)}")
        return redirect('my-programs')
    

import boto3
from botocore.exceptions import ClientError
from django.conf import settings
from django.contrib import messages
from django.http import JsonResponse
import environ
import os
from concurrent.futures import ThreadPoolExecutor

env = environ.Env()
environ.Env.read_env()  # This loads variables from your .env file

def export_and_copy_to_s3(request, schedule_date):
    """
    Export playlist and copy to S3 bucket
    Returns JSON response with success/error status
    """
    try:
        # First generate the playlist file using existing export function
        response = export_playlist(schedule_date)
        
        # Get the file content from the response
        playlist_content = response.content.decode('utf-8')
        filename = response.get('Content-Disposition').split('filename=')[1].strip('"')
        
        # Initialize S3 client
        s3_client = boto3.client('s3',
            aws_access_key_id = env('AWS_ACCESS_KEY_ID', default=None),
            aws_secret_access_key = env('AWS_SECRET_ACCESS_KEY', default=None),
            region_name = env('AWS_REGION', default='us-east-1')  # Optional default region
        )

        # S3 bucket and key
        bucket = 'channel24-playlist-9cf77ba1-577f-4924-80c0-3f0c12c6ac07'
        s3_key = filename
        
        # Upload to S3
        try:
            s3_client.put_object(
                Bucket=bucket,
                Key=s3_key,
                Body=playlist_content.encode('utf-8'),
                ContentType='text/plain'
            )

            # Initialize Storage Gateway client
            sgw_client = boto3.client('storagegateway',
                aws_access_key_id = env('AWS_ACCESS_KEY_ID', default=None),
                aws_secret_access_key = env('AWS_SECRET_ACCESS_KEY', default=None),
                region_name = env('AWS_REGION', default='us-east-1')  # Optional default region
            )
            
            # Function to refresh a single file share
            def refresh_file_share(file_share_id):
                try:
                    response = sgw_client.refresh_cache(
                        FileShareARN=file_share_id,
                        FolderList=['/']  # Refresh the entire share
                    )
                    return {
                        'file_share_id': file_share_id,
                        'status': 'success',
                        'refresh_job_id': response.get('RefreshCacheJobId')
                    }
                except Exception as e:
                    return {
                        'file_share_id': file_share_id,
                        'status': 'error',
                        'error': str(e)
                    }
            
            # Get file shares from settings and split into list
            storage_gateway_file_shares_str = env('STORAGE_GATEWAY_FILE_SHARES_STR', default=None)
            file_shares = storage_gateway_file_shares_str.split(',')
            
            # Use ThreadPoolExecutor to refresh cache in parallel
            refresh_results = []
            with ThreadPoolExecutor(max_workers=len(file_shares)) as executor:
                refresh_results = list(executor.map(refresh_file_share, file_shares))
            
            # Check if any refreshes failed
            failed_refreshes = [r for r in refresh_results if r['status'] == 'error']
            
            if failed_refreshes:
                return JsonResponse({
                    'status': 'partial_success',
                    'message': f'Playlist exported to S3, but some cache refreshes failed',
                    'filename': filename,
                    'refresh_results': refresh_results
                })
            
            return JsonResponse({
                'status': 'success',
                'message': f'Playlist exported to S3 and cache refreshed successfully',
                'filename': filename,
                'refresh_results': refresh_results
            })
            
        except ClientError as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Failed to upload to S3: {str(e)}'
            }, status=500)
            
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'Error processing request: {str(e)}'
        }, status=500)
    

import http.client as http_client
import logging
import boto3
from django.http import HttpResponse
from django.core.mail import send_mail

# Enable low-level HTTP debug logging for the underlying HTTP connection.
http_client.HTTPConnection.debuglevel = 1

# Configure logging to show debug output.
logging.basicConfig(level=logging.DEBUG)
logging.getLogger('botocore').setLevel(logging.DEBUG)
logging.getLogger('urllib3').setLevel(logging.DEBUG)
logging.getLogger('django_ses').setLevel(logging.DEBUG)

def test_email(request):
    logger = logging.getLogger('django.core.mail')
    logger.info("Sending test email with detailed HTTP debug logging...")
    result = send_mail(
       'Test Email from Django and SES',
        'This is a test email to see the full protocol exchange details.',
        'no-reply@atlanta24communitymedia.com',
        ['jpound@AtlantaGa.Gov'],
        fail_silently=False
    )
    logger.info("Emails sent: %s", result)
    return HttpResponse("Test email sent!")

import os
from django.conf import settings
from django.http import HttpResponse, Http404

def acme_challenge_view(request, token):
    # Define the directory where your challenge files are stored.
    # You could also put these files in a specific directory like
    # os.path.join(settings.BASE_DIR, 'acme-challenge')
    challenge_dir = os.path.join(settings.BASE_DIR, 'acme-challenge')
    file_path = os.path.join(challenge_dir, token)
    
    if os.path.exists(file_path):
        with open(file_path, 'rb') as f:
            return HttpResponse(f.read(), content_type='text/plain')
    else:
        raise Http404("ACME challenge file not found.")

# views.py
from django.shortcuts import render, get_object_or_404, redirect
from .models import Program, Creator, TIME_SLOTS_CHOICES  # Add TIME_SLOTS_CHOICES import
from .forms import ProgramOverrideForm
from .models import GENRE_CHOICES  # if defined in models or import from your constants

def is_admin(user):
    return user.is_staff or user.is_superuser

@login_required
@user_passes_test(is_admin)
def available_programs(request):
    # Get filter parameters from GET
    search = request.GET.get('search', '')
    creator_filter = request.GET.get('creator', '')
    genre_filter = request.GET.get('genre', '')

    # Start with all programs
    programs = Program.objects.all()

    # Apply filters
    if search:
        programs = programs.filter(program_name__icontains=search)
    if creator_filter:
        programs = programs.filter(creator__custom_id=creator_filter)
    if genre_filter:
        programs = programs.filter(genre=genre_filter)

    # Handle update submissions
    if request.method == 'POST':
        program_id = request.POST.get('program_id')
        program = get_object_or_404(Program, custom_id=program_id)
        form = ProgramOverrideForm(request.POST, instance=program)
        if form.is_valid():
            form.save()
            # Optionally add a success message here
            return redirect('available_programs')

    # Build a list of tuples: each program with its own override form instance
    programs_with_forms = [
        (program, ProgramOverrideForm(instance=program))
        for program in programs
    ]
    
    # Debug - print each form's time_slots field
    # for program, form in programs_with_forms:
    #     print(f"Program: {program.program_name}")
    #     print(f"Form time_slots field: {form.fields['override_time_slots']}")
    #     print(f"Form time_slots field type: {type(form.fields['override_time_slots'])}")
    #     print(f"Form time_slots dir: {dir(form.fields['override_time_slots'])}")
    
    context = {
        'programs_with_forms': programs_with_forms,
        'creators': Creator.objects.all(),  
        'GENRE_CHOICES': GENRE_CHOICES,
        'TIME_SLOTS_CHOICES': TIME_SLOTS_CHOICES,  # Add this line
        'current_filters': {
            'search': search,
            'creator': creator_filter,
            'genre': genre_filter,
        },
    }
    return render(request, 'available_programs.html', context)


@login_required
def my_schedule(request):
    logger.debug("Entered my_schedule view for user: %s", request.user.username)

    # Get all creators associated with the logged-in user
    creators = Creator.objects.filter(created_by=request.user)
    
    if not creators.exists():
        logger.warning("No creators found for user: %s", request.user.username)
        messages.error(request, "You don't have any creator profiles.")
        return render(request, 'my_schedule.html', {'scheduled_episodes': [], 'creators': []})
    
    # Get the selected date from the request or use today's date
    selected_date_str = request.GET.get('schedule_date')
    if selected_date_str:
        try:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except ValueError:
            # Fallback if parsing fails
            selected_date = timezone.now().date()
            logger.warning("Invalid date format received: %s. Using today's date.", selected_date_str)
    else:
        selected_date = timezone.now().date()
    
    logger.debug("Selected date for schedule: %s", selected_date)

    # Get a specific creator if provided in the request
    selected_creator_id = request.GET.get('creator')
    
    # If a specific creator is selected, filter by that creator
    if selected_creator_id:
        try:
            selected_creator = creators.get(custom_id=selected_creator_id)
            creator_filter = [selected_creator]
        except (Creator.DoesNotExist, ValueError):
            # If the selected creator doesn't exist or isn't valid, use all creators
            creator_filter = creators
    else:
        # Use all creators if none is specifically selected
        creator_filter = creators

    # Query the schedule for the creator(s) using the selected date
    scheduled_episodes = ScheduledEpisode.objects.filter(
        schedule_date=selected_date,
        creator__in=creator_filter  # Filter by the user's creators
    ).select_related('episode', 'program', 'creator').order_by('start_time')

    context = {
        'scheduled_episodes': scheduled_episodes,
        'selected_date': selected_date.strftime('%Y-%m-%d'),
        'creators': creators,
        'selected_creator_id': selected_creator_id
    }

    logger.debug(
        "Rendering my_schedule.html with %s scheduled episodes for date: %s",
        scheduled_episodes.count(),
        selected_date
    )

    return render(request, 'my_schedule.html', context)

# from django.shortcuts import render
# from django.views.generic import ListView
# from django.db.models import Q
# from datetime import datetime
# from .models import Episode, Creator, Program

# class ContentReportView(ListView):
#     model = Episode
#     template_name = 'content_report.html'
#     context_object_name = 'episodes'
#     paginate_by = 20

#     def get_queryset(self):
#         # queryset = Episode.objects.select_related('program', 'program__creator').all()
#         queryset = Episode.objects.select_related('program', 'program__creator').exclude(
#             Q(duration_timecode__isnull=True) |
#             Q(duration_timecode__exact='') |
#             Q(duration_timecode__exact='None') |
#             Q(duration_timecode__exact='null')
#         )
        
#         # Handle search filters
#         channel_name = self.request.GET.get('channel_name', '')
#         program_name = self.request.GET.get('program_name', '')
#         start_date = self.request.GET.get('start_date', '')
#         end_date = self.request.GET.get('end_date', '')
#         search_query = self.request.GET.get('search', '')
        
#         # Apply filters
#         if channel_name:
#             queryset = queryset.filter(program__creator__channel_name__icontains=channel_name)
        
#         if program_name:
#             queryset = queryset.filter(program__program_name__icontains=program_name)
        
#         if start_date and end_date:
#             try:
#                 start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
#                 end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
#                 queryset = queryset.filter(created_at__range=(start_date_obj, end_date_obj))
#             except ValueError:
#                 pass
#         elif start_date:
#             try:
#                 start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
#                 queryset = queryset.filter(created_at__gte=start_date_obj)
#             except ValueError:
#                 pass
#         elif end_date:
#             try:
#                 end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
#                 queryset = queryset.filter(created_at__lte=end_date_obj)
#             except ValueError:
#                 pass
                
#         # General search across multiple fields
#         if search_query:
#             queryset = queryset.filter(
#                 Q(program__creator__channel_name__icontains=search_query) |
#                 Q(program__program_name__icontains=search_query) |
#                 Q(title__icontains=search_query) |
#                 Q(description__icontains=search_query) |
#                 Q(ai_summary__icontains=search_query) |
#                 Q(ai_topics__contains=[search_query])
#             )
        
#         # Default sorting by most recent
#         sort_by = self.request.GET.get('sort', '-created_at')
#         return queryset.order_by(sort_by)
    
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
        
#         # Add creators and programs for filter dropdowns
#         context['creators'] = Creator.objects.all()
#         context['programs'] = Program.objects.all()
        
#         # Add current filter values for form
#         context['current_filters'] = {
#             'channel_name': self.request.GET.get('channel_name', ''),
#             'program_name': self.request.GET.get('program_name', ''),
#             'start_date': self.request.GET.get('start_date', ''),
#             'end_date': self.request.GET.get('end_date', ''),
#             'search': self.request.GET.get('search', ''),
#             'sort': self.request.GET.get('sort', '-created_at')
#         }
        
#         # Add sort options
#         context['sort_options'] = [
#             ('-created_at', 'Newest First'),
#             ('created_at', 'Oldest First'),
#             ('program__creator__channel_name', 'Channel Name (A-Z)'),
#             ('-program__creator__channel_name', 'Channel Name (Z-A)'),
#             ('program__program_name', 'Program Name (A-Z)'),
#             ('-program__program_name', 'Program Name (Z-A)'),
#             ('episode_number', 'Episode Number (Ascending)'),
#             ('-episode_number', 'Episode Number (Descending)'),
#             ('title', 'Title (A-Z)'),
#             ('-title', 'Title (Z-A)'),
#             ('schedule_count', 'Schedule Count (Ascending)'),
#             ('-schedule_count', 'Schedule Count (Descending)'),
#             ('last_scheduled', 'Last Scheduled (Ascending)'),
#             ('-last_scheduled', 'Last Scheduled (Descending)'),
#         ]
        
#         return context

from django.shortcuts import render
from django.views.generic import ListView
from django.db.models import Q
from django.http import HttpResponse
from datetime import datetime
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from .models import Episode, Creator, Program

class ContentReportView(ListView):
    model = Episode
    template_name = 'content_report.html'
    context_object_name = 'episodes'
    paginate_by = 20

    def get_queryset(self):
        # Start with episodes that have valid duration - exclude None, empty, and 'None' string values
        queryset = Episode.objects.select_related('program', 'program__creator').exclude(
            Q(duration_timecode__isnull=True) |
            Q(duration_timecode__exact='') |
            Q(duration_timecode__exact='None') |
            Q(duration_timecode__exact='null')
        )
        
        # Handle search filters
        channel_name = self.request.GET.get('channel_name', '')
        program_name = self.request.GET.get('program_name', '')
        start_date = self.request.GET.get('start_date', '')
        end_date = self.request.GET.get('end_date', '')
        search_query = self.request.GET.get('search', '')
        
        # Apply filters
        if channel_name:
            queryset = queryset.filter(program__creator__channel_name__icontains=channel_name)
        
        if program_name:
            queryset = queryset.filter(program__program_name__icontains=program_name)
        
        if start_date and end_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
                queryset = queryset.filter(created_at__range=(start_date_obj, end_date_obj))
            except ValueError:
                pass
        elif start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
                queryset = queryset.filter(created_at__gte=start_date_obj)
            except ValueError:
                pass
        elif end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
                queryset = queryset.filter(created_at__lte=end_date_obj)
            except ValueError:
                pass
                
        # General search across multiple fields
        if search_query:
            queryset = queryset.filter(
                Q(program__creator__channel_name__icontains=search_query) |
                Q(program__program_name__icontains=search_query) |
                Q(title__icontains=search_query) |
                Q(description__icontains=search_query) |
                Q(ai_summary__icontains=search_query) |
                Q(ai_topics__contains=[search_query])
            )
        
        # Default sorting by most recent
        sort_by = self.request.GET.get('sort', '-created_at')
        return queryset.order_by(sort_by)

    def get(self, request, *args, **kwargs):
        # Check if this is an export request
        export_format = request.GET.get('export')
        if export_format in ['csv', 'excel']:
            return self.export_data(export_format)
        
        # Otherwise, return normal ListView response
        return super().get(request, *args, **kwargs)

    def export_data(self, format_type):
        # Get the filtered queryset (same as get_queryset but without pagination)
        queryset = self.get_queryset()
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format_type == 'csv':
            return self.export_csv(queryset, timestamp)
        elif format_type == 'excel':
            return self.export_excel(queryset, timestamp)

    def export_csv(self, queryset, timestamp):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="content_report_{timestamp}.csv"'
        
        writer = csv.writer(response)
        
        # Write header row
        writer.writerow([
            'Channel Name',
            'Creator First Name',
            'Creator Last Name',
            'Program Name',
            'Episode Number',
            'Title',
            'Description',
            'Date Uploaded',
            'Age Rating',
            'Duration',
            'Schedule Count',
            'Last Scheduled',
            'AI Summary',
            'AI Topics'
        ])
        
        # Write data rows
        for episode in queryset:
            writer.writerow([
                episode.program.creator.channel_name,
                episode.program.creator.first_name,
                episode.program.creator.last_name,
                episode.program.program_name,
                episode.episode_number,
                episode.title,
                episode.description,
                episode.created_at.strftime('%Y-%m-%d') if episode.created_at else '',
                episode.ai_age_rating,
                episode.duration_timecode,
                episode.schedule_count,
                episode.last_scheduled.strftime('%Y-%m-%d') if episode.last_scheduled else 'Never',
                episode.ai_summary,
                ', '.join(episode.ai_topics) if episode.ai_topics else ''
            ])
        
        return response

    def export_excel(self, queryset, timestamp):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="content_report_{timestamp}.xlsx"'
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Content Report'
        
        # Define headers
        headers = [
            'Channel Name',
            'Creator First Name',
            'Creator Last Name',
            'Program Name',
            'Episode Number',
            'Title',
            'Description',
            'Date Uploaded',
            'Age Rating',
            'Duration',
            'Schedule Count',
            'Last Scheduled',
            'AI Summary',
            'AI Topics'
        ]
        
        # Write headers with styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write data rows
        for row_num, episode in enumerate(queryset, 2):
            worksheet.cell(row=row_num, column=1, value=episode.program.creator.channel_name)
            worksheet.cell(row=row_num, column=2, value=episode.program.creator.first_name)
            worksheet.cell(row=row_num, column=3, value=episode.program.creator.last_name)
            worksheet.cell(row=row_num, column=4, value=episode.program.program_name)
            worksheet.cell(row=row_num, column=5, value=episode.episode_number)
            worksheet.cell(row=row_num, column=6, value=episode.title)
            worksheet.cell(row=row_num, column=7, value=episode.description)
            worksheet.cell(row=row_num, column=8, value=episode.created_at.strftime('%Y-%m-%d') if episode.created_at else '')
            worksheet.cell(row=row_num, column=9, value=episode.ai_age_rating)
            worksheet.cell(row=row_num, column=10, value=episode.duration_timecode)
            worksheet.cell(row=row_num, column=11, value=episode.schedule_count)
            worksheet.cell(row=row_num, column=12, value=episode.last_scheduled.strftime('%Y-%m-%d') if episode.last_scheduled else 'Never')
            worksheet.cell(row=row_num, column=13, value=episode.ai_summary)
            worksheet.cell(row=row_num, column=14, value=', '.join(episode.ai_topics) if episode.ai_topics else '')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(response)
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add creators and programs for filter dropdowns
        context['creators'] = Creator.objects.all()
        context['programs'] = Program.objects.all()
        
        # Add current filter values for form
        context['current_filters'] = {
            'channel_name': self.request.GET.get('channel_name', ''),
            'program_name': self.request.GET.get('program_name', ''),
            'start_date': self.request.GET.get('start_date', ''),
            'end_date': self.request.GET.get('end_date', ''),
            'search': self.request.GET.get('search', ''),
            'sort': self.request.GET.get('sort', '-created_at')
        }
        
        # Add sort options
        context['sort_options'] = [
            ('-created_at', 'Newest First'),
            ('created_at', 'Oldest First'),
            ('program__creator__channel_name', 'Channel Name (A-Z)'),
            ('-program__creator__channel_name', 'Channel Name (Z-A)'),
            ('program__program_name', 'Program Name (A-Z)'),
            ('-program__program_name', 'Program Name (Z-A)'),
            ('episode_number', 'Episode Number (Ascending)'),
            ('-episode_number', 'Episode Number (Descending)'),
            ('title', 'Title (A-Z)'),
            ('-title', 'Title (Z-A)'),
            ('schedule_count', 'Schedule Count (Ascending)'),
            ('-schedule_count', 'Schedule Count (Descending)'),
            ('last_scheduled', 'Last Scheduled (Ascending)'),
            ('-last_scheduled', 'Last Scheduled (Descending)'),
        ]
        
        return context

from django.shortcuts import render
from django.views.generic import ListView
from django.db.models import Q, Min
from django.http import HttpResponse
from datetime import datetime, timedelta
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from .models import ScheduledEpisode, Creator, Program

class ScheduleReportView(ListView):
    model = ScheduledEpisode
    template_name = 'schedule_report.html'
    context_object_name = 'scheduled_episodes'
    paginate_by = 20

    def get_queryset(self):
        from django.db.models import Window, F
        from django.db.models.functions import FirstValue
        
        # Start with scheduled episodes, select related data to avoid N+1 queries
        queryset = ScheduledEpisode.objects.select_related(
            'episode', 'program', 'creator'
        ).all()
        
        # Annotate with first scheduled date for each episode
        queryset = queryset.annotate(
            first_scheduled_date=Window(
                expression=FirstValue('schedule_date'),
                partition_by=[F('episode')],
                order_by=F('schedule_date').asc()
            )
        )
        
        # Handle search filters - strip whitespace from all inputs
        channel_name = self.request.GET.get('channel_name', '').strip()
        program_name = self.request.GET.get('program_name', '').strip()
        start_date = self.request.GET.get('start_date', '').strip()
        end_date = self.request.GET.get('end_date', '').strip()
        search_query = self.request.GET.get('search', '').strip()
        status_filter = self.request.GET.get('status', '').strip()
        
        # Apply filters
        if channel_name:
            queryset = queryset.filter(creator__channel_name__icontains=channel_name)
        
        if program_name:
            queryset = queryset.filter(program__program_name__icontains=program_name)
        
        if start_date and end_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                queryset = queryset.filter(schedule_date__range=(start_date_obj, end_date_obj))
            except ValueError as e:
                pass
        elif start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                queryset = queryset.filter(schedule_date__gte=start_date_obj)
            except ValueError as e:
                pass
        elif end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                queryset = queryset.filter(schedule_date__lte=end_date_obj)
            except ValueError as e:
                pass
        
        # Status filter (New/Repeat) - now using the annotation
        if status_filter:
            if status_filter == 'new':
                # Filter for episodes where schedule_date equals first_scheduled_date
                queryset = queryset.filter(schedule_date=F('first_scheduled_date'))
                
            elif status_filter == 'repeat':
                # Filter for episodes where schedule_date is after first_scheduled_date
                queryset = queryset.filter(schedule_date__gt=F('first_scheduled_date'))
                
        # General search across multiple fields
        if search_query:
            queryset = queryset.filter(
                Q(creator__channel_name__icontains=search_query) |
                Q(program__program_name__icontains=search_query) |
                Q(title__icontains=search_query) |
                Q(episode__description__icontains=search_query) |
                Q(ai_topics__contains=[search_query])
            )
        
        # Default sorting by schedule date and time
        sort_by = self.request.GET.get('sort', 'schedule_date')
        return queryset.order_by(sort_by, 'start_time')

    def get_program_status(self, scheduled_episode):
        """
        Determine if a program is 'New' or 'Repeat' based on:
        Whether this is the first DATE this episode was ever scheduled
        """
        if not scheduled_episode.episode:
            return 'Unknown'
        
        # Check if we have the annotation from get_queryset
        if hasattr(scheduled_episode, 'first_scheduled_date'):
            if scheduled_episode.schedule_date == scheduled_episode.first_scheduled_date:
                return 'New'
            else:
                return 'Repeat'
        
        # Fallback: Find the earliest date this episode was scheduled
        first_scheduled_date = ScheduledEpisode.objects.filter(
            episode=scheduled_episode.episode
        ).aggregate(
            first_date=Min('schedule_date')
        )['first_date']
        
        if first_scheduled_date:
            # If this scheduled episode's date matches the first scheduled date, it's "New"
            if scheduled_episode.schedule_date == first_scheduled_date:
                return 'New'
            else:
                return 'Repeat'
        
        # Final fallback: if no scheduled date found, check if episode has been scheduled at all
        if scheduled_episode.episode.schedule_count is None or scheduled_episode.episode.schedule_count == 0:
            return 'New'
        
        return 'Unknown'

    def get(self, request, *args, **kwargs):
        # Check if this is an export request
        export_format = request.GET.get('export')
        if export_format == 'csv':
            return self.export_csv()
        elif export_format == 'excel':
            return self.export_excel()
        
        # Otherwise, return normal ListView response
        return super().get(request, *args, **kwargs)

    def export_csv(self):
        # Get the filtered queryset (same as get_queryset but without pagination)
        queryset = self.get_queryset()
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="schedule_report_{timestamp}.csv"'
        
        writer = csv.writer(response)
        
        # Write header row
        writer.writerow([
            'Air Date',
            'Air Time',
            'Program Name',
            'Episode Number',
            'Episode Title',
            'Producer Name',
            'Channel Name',
            'Duration',
            'Program Status',
            'Age Rating',
            'Genre',
            'Topics'
        ])
        
        # Write data rows
        for scheduled_episode in queryset:
            writer.writerow([
                scheduled_episode.schedule_date.strftime('%Y-%m-%d') if scheduled_episode.schedule_date else '',
                scheduled_episode.start_time.strftime('%H:%M:%S') if scheduled_episode.start_time else '',
                scheduled_episode.program.program_name,
                scheduled_episode.episode_number,
                scheduled_episode.title,
                f"{scheduled_episode.creator.first_name} {scheduled_episode.creator.last_name}",
                scheduled_episode.creator.channel_name,
                scheduled_episode.duration_timecode or '',
                self.get_program_status(scheduled_episode),
                scheduled_episode.ai_age_rating or '',
                scheduled_episode.ai_genre or '',
                ', '.join(scheduled_episode.ai_topics) if scheduled_episode.ai_topics else ''
            ])
        
        return response

    def export_excel(self):
        # Get the filtered queryset (same as get_queryset but without pagination)
        queryset = self.get_queryset()
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="schedule_report_{timestamp}.xlsx"'
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Schedule Report'
        
        # Define headers
        headers = [
            'Air Date',
            'Air Time',
            'Program Name',
            'Episode Number',
            'Episode Title',
            'Producer Name',
            'Channel Name',
            'Duration',
            'Program Status',
            'Age Rating',
            'Genre',
            'Topics'
        ]
        
        # Write headers with styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write data rows
        for row_num, scheduled_episode in enumerate(queryset, 2):
            worksheet.cell(row=row_num, column=1, value=scheduled_episode.schedule_date.strftime('%Y-%m-%d') if scheduled_episode.schedule_date else '')
            worksheet.cell(row=row_num, column=2, value=scheduled_episode.start_time.strftime('%H:%M:%S') if scheduled_episode.start_time else '')
            worksheet.cell(row=row_num, column=3, value=scheduled_episode.program.program_name)
            worksheet.cell(row=row_num, column=4, value=scheduled_episode.episode_number)
            worksheet.cell(row=row_num, column=5, value=scheduled_episode.title)
            worksheet.cell(row=row_num, column=6, value=f"{scheduled_episode.creator.first_name} {scheduled_episode.creator.last_name}")
            worksheet.cell(row=row_num, column=7, value=scheduled_episode.creator.channel_name)
            worksheet.cell(row=row_num, column=8, value=scheduled_episode.duration_timecode or '')
            worksheet.cell(row=row_num, column=9, value=self.get_program_status(scheduled_episode))
            worksheet.cell(row=row_num, column=10, value=scheduled_episode.ai_age_rating or '')
            worksheet.cell(row=row_num, column=11, value=scheduled_episode.ai_genre or '')
            worksheet.cell(row=row_num, column=12, value=', '.join(scheduled_episode.ai_topics) if scheduled_episode.ai_topics else '')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        workbook.save(response)
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add creators and programs for filter dropdowns
        context['creators'] = Creator.objects.all()
        context['programs'] = Program.objects.all()
        
        # Add current filter values for form - strip whitespace
        context['current_filters'] = {
            'channel_name': self.request.GET.get('channel_name', '').strip(),
            'program_name': self.request.GET.get('program_name', '').strip(),
            'start_date': self.request.GET.get('start_date', '').strip(),
            'end_date': self.request.GET.get('end_date', '').strip(),
            'search': self.request.GET.get('search', '').strip(),
            'status': self.request.GET.get('status', '').strip(),
            'sort': self.request.GET.get('sort', 'schedule_date').strip()
        }
        
        # Add sort options
        context['sort_options'] = [
            ('schedule_date', 'Air Date (Earliest First)'),
            ('-schedule_date', 'Air Date (Latest First)'),
            ('start_time', 'Air Time (Earliest First)'),
            ('-start_time', 'Air Time (Latest First)'),
            ('creator__channel_name', 'Channel Name (A-Z)'),
            ('-creator__channel_name', 'Channel Name (Z-A)'),
            ('program__program_name', 'Program Name (A-Z)'),
            ('-program__program_name', 'Program Name (Z-A)'),
            ('episode_number', 'Episode Number (Ascending)'),
            ('-episode_number', 'Episode Number (Descending)'),
            ('title', 'Title (A-Z)'),
            ('-title', 'Title (Z-A)'),
        ]
        
        # Add status options for filter
        context['status_options'] = [
            ('', 'All Status'),
            ('new', 'New'),
            ('repeat', 'Repeat'),
        ]
        
        # Add method to determine program status for template
        context['get_program_status'] = self.get_program_status
        
        return context